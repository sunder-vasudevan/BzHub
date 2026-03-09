"""Inventory management tab — CRUD, search, CSV import/export."""
import csv
import os
import shutil
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

from src.core import CurrencyFormatter
from .base_tab import BaseTab


class InventoryTab(BaseTab):
    """Full inventory management: add, update, delete, search, import/export."""

    def __init__(self, notebook: ttk.Notebook, app, tab_label: str = "📦 Inventory"):
        super().__init__(notebook, app)
        self._tab_label = tab_label
        self._build()

    # ------------------------------------------------------------------
    # Build UI
    # ------------------------------------------------------------------

    def _build(self):
        self.notebook.add(self.frame, text=self._tab_label)

        container = tk.Frame(self.frame, bg=self.colors["bg"])
        container.pack(fill="both", expand=True, padx=12, pady=12)

        # Header
        header = tk.Frame(container, bg=self.colors["bg"])
        header.pack(fill="x", pady=(0, 12))
        ttk.Label(header, text="📦 Inventory", style="Header.TLabel").pack(side="left")
        ttk.Label(header, text="Manage stock, pricing, and thresholds",
                  style="Subheader.TLabel").pack(side="left", padx=10)

        body = tk.Frame(container, bg=self.colors["bg"])
        body.pack(fill="both", expand=True)

        # Left: form card
        form_card = tk.Frame(body, bg=self.colors["card"], padx=12, pady=12)
        form_card.pack(side="left", fill="y", padx=(0, 10))

        tk.Label(form_card, text="Add / Edit Item", bg=self.colors["card"],
                 fg=self.colors["text"], font=("Arial", 10, "bold")).pack(anchor="w", pady=(0, 8))

        def field(label_text):
            row = tk.Frame(form_card, bg=self.colors["card"])
            row.pack(fill="x", pady=4)
            tk.Label(row, text=label_text, bg=self.colors["card"],
                     fg=self.colors["muted"], width=12, anchor="w").pack(side="left")
            entry = ttk.Entry(row)
            entry.pack(side="left", fill="x", expand=True)
            return entry

        self._name     = field("Item Name")
        self._qty      = field("Quantity")
        self._threshold = field("Threshold")
        self._cost     = field("Cost Price")
        self._sale     = field("Sale Price")
        self._desc     = field("Description")

        # Image row with browse button
        img_row = tk.Frame(form_card, bg=self.colors["card"])
        img_row.pack(fill="x", pady=4)
        tk.Label(img_row, text="Image", bg=self.colors["card"],
                 fg=self.colors["muted"], width=12, anchor="w").pack(side="left")
        self._image = ttk.Entry(img_row)
        self._image.pack(side="left", fill="x", expand=True)
        ttk.Button(img_row, text="Browse", style="Info.TButton",
                   command=self._browse_image).pack(side="left", padx=4)

        btn_row = tk.Frame(form_card, bg=self.colors["card"])
        btn_row.pack(fill="x", pady=(10, 0))
        ttk.Button(btn_row, text="Add",    command=self._add,    style="Success.TButton").pack(side="left", padx=3)
        ttk.Button(btn_row, text="Update", command=self._update, style="Primary.TButton").pack(side="left", padx=3)
        ttk.Button(btn_row, text="Delete", command=self._delete, style="Danger.TButton").pack(side="left", padx=3)
        ttk.Button(btn_row, text="Clear",  command=self._clear,  style="Info.TButton").pack(side="left", padx=3)

        # Right: list card
        list_card = tk.Frame(body, bg=self.colors["card"], padx=12, pady=12)
        list_card.pack(side="left", fill="both", expand=True)

        list_header = tk.Frame(list_card, bg=self.colors["card"])
        list_header.pack(fill="x")
        tk.Label(list_header, text="Inventory Items", bg=self.colors["card"],
                 fg=self.colors["text"], font=("Arial", 10, "bold")).pack(side="left")

        # Import/Export toolbar
        io_row = tk.Frame(list_card, bg=self.colors["card"])
        io_row.pack(fill="x", pady=(4, 8))
        ttk.Button(io_row, text="Export CSV",   command=self._export_csv,   style="Info.TButton").pack(side="left", padx=2)
        ttk.Button(io_row, text="Export Excel", command=self._export_excel, style="Info.TButton").pack(side="left", padx=2)
        ttk.Button(io_row, text="Import CSV",   command=self._import_csv,   style="Info.TButton").pack(side="left", padx=12)
        ttk.Button(io_row, text="Import Excel", command=self._import_excel, style="Info.TButton").pack(side="left", padx=2)

        # Search row
        search_row = tk.Frame(list_card, bg=self.colors["card"])
        search_row.pack(fill="x", pady=(0, 6))
        tk.Label(search_row, text="Search", bg=self.colors["card"],
                 fg=self.colors["muted"]).pack(side="left", padx=(0, 6))
        self._search = ttk.Entry(search_row)
        self._search.pack(side="left", fill="x", expand=True)
        ttk.Button(search_row, text="🔎",     command=self._search_items,   style="Info.TButton").pack(side="left", padx=6)
        ttk.Button(search_row, text="Refresh", command=self.refresh,         style="Info.TButton").pack(side="left")

        # Inventory treeview
        cols = ("Item", "Qty", "Threshold", "Cost", "Sale", "Description")
        self._tree = ttk.Treeview(list_card, columns=cols, height=18, show="headings")
        for col, width in [("Item", 180), ("Qty", 70), ("Threshold", 90),
                           ("Cost", 90), ("Sale", 90), ("Description", 220)]:
            self._tree.column(col, anchor="center", width=width)
            self._tree.heading(col, text=col)
        self._tree.pack(fill="both", expand=True, pady=(6, 0))
        self._tree.bind("<Double-1>", self._on_select)

        self.refresh()

    # ------------------------------------------------------------------
    # Data refresh
    # ------------------------------------------------------------------

    def refresh(self):
        self._load_items(self.app.inventory_service.get_all_items())

    def _load_items(self, items):
        for row in self._tree.get_children():
            self._tree.delete(row)
        for item in items:
            self._tree.insert("", "end", values=(
                item[0], item[1], item[2],
                CurrencyFormatter.format_currency(item[3]),
                CurrencyFormatter.format_currency(item[4]),
                item[5] or "",
            ))

    # ------------------------------------------------------------------
    # Form actions
    # ------------------------------------------------------------------

    def _add(self):
        try:
            name  = self._name.get().strip()
            qty   = int(self._qty.get())
            thr   = int(self._threshold.get())
            cost  = float(self._cost.get())
            sale  = float(self._sale.get())
            desc  = self._desc.get().strip()
            image = self._image.get().strip() or None

            if not name:
                messagebox.showerror("Error", "Item name required")
                return
            if self.app.inventory_service.add_item(name, qty, thr, cost, sale, desc,
                                                   image_path=image):
                self.app.activity_service.log(
                    self.app.current_user, "Add Inventory", f"Added item: {name}")
                messagebox.showinfo("Success", f"Item '{name}' added")
                self._clear()
                self.refresh()
            else:
                messagebox.showerror("Error", "Failed to add item (may already exist)")
        except ValueError as e:
            messagebox.showerror("Error", f"Invalid input: {e}")

    def _update(self):
        try:
            name = self._name.get().strip()
            if not name:
                messagebox.showerror("Error", "Select item to update")
                return
            qty  = int(self._qty.get())       if self._qty.get()       else None
            thr  = int(self._threshold.get()) if self._threshold.get() else None
            cost = float(self._cost.get())    if self._cost.get()      else None
            sale = float(self._sale.get())    if self._sale.get()      else None
            desc = self._desc.get().strip()   or None

            if self.app.inventory_service.update_item(
                name, quantity=qty, threshold=thr,
                cost_price=cost, sale_price=sale, description=desc,
            ):
                self.app.activity_service.log(
                    self.app.current_user, "Update Inventory", f"Updated item: {name}")
                messagebox.showinfo("Success", f"Item '{name}' updated")
                self.refresh()
            else:
                messagebox.showerror("Error", "Failed to update item")
        except Exception as e:
            messagebox.showerror("Error", f"Invalid input: {e}")

    def _delete(self):
        try:
            name = self._name.get().strip()
            if not name:
                messagebox.showerror("Error", "Select item to delete")
                return
            messagebox.showwarning("Caution", "This will permanently delete the inventory item.")
            if messagebox.askyesno("Confirm", f"Delete '{name}'?"):
                if self.app.inventory_service.delete_item(name):
                    self.app.activity_service.log(
                        self.app.current_user, "Delete Inventory", f"Deleted item: {name}")
                    messagebox.showinfo("Success", "Item deleted")
                    self._clear()
                    self.refresh()
                else:
                    messagebox.showerror("Error", "Failed to delete item")
        except Exception as e:
            messagebox.showerror("Error", f"Error deleting item: {e}")

    def _clear(self):
        for entry in [self._name, self._qty, self._threshold,
                      self._cost, self._sale, self._desc, self._image]:
            entry.delete(0, tk.END)

    def _search_items(self):
        query = self._search.get().strip()
        items = (self.app.inventory_service.search(query) if query
                 else self.app.inventory_service.get_all_items())
        self._load_items(items)

    def _on_select(self, event=None):
        sel = self._tree.selection()
        if not sel:
            return
        vals = self._tree.item(sel[0])["values"]
        self._name.delete(0, tk.END)
        self._name.insert(0, vals[0])
        self._qty.delete(0, tk.END)
        self._qty.insert(0, vals[1])
        self._threshold.delete(0, tk.END)
        self._threshold.insert(0, vals[2])
        self._cost.delete(0, tk.END)
        self._cost.insert(0, CurrencyFormatter.parse_currency(str(vals[3])))
        self._sale.delete(0, tk.END)
        self._sale.insert(0, CurrencyFormatter.parse_currency(str(vals[4])))
        self._desc.delete(0, tk.END)
        self._desc.insert(0, vals[5])
        details = self.app.inventory_service.get_item(vals[0])
        self._image.delete(0, tk.END)
        if details and details.get("image_path"):
            self._image.insert(0, details["image_path"])

    def _browse_image(self):
        path = filedialog.askopenfilename(
            title="Select Item Image",
            filetypes=[("Image Files", "*.png *.jpg *.jpeg *.gif *.bmp"),
                       ("All Files", "*.*")],
        )
        if not path:
            return
        try:
            dest_dir = os.path.abspath(
                os.path.join(os.path.dirname(__file__), "../../../../assets/item_images")
            )
            os.makedirs(dest_dir, exist_ok=True)
            dest = os.path.join(dest_dir, os.path.basename(path))
            if os.path.abspath(path) != os.path.abspath(dest):
                shutil.copy(path, dest)
            self._image.delete(0, tk.END)
            self._image.insert(0, dest)
        except Exception as e:
            messagebox.showerror("Image", f"Failed to attach image: {e}")

    # ------------------------------------------------------------------
    # Export / Import
    # ------------------------------------------------------------------

    def _export_csv(self):
        items = self.app.inventory_service.get_all_items()
        if not items:
            messagebox.showwarning("Export", "No inventory data to export.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")],
            title="Export Inventory to CSV",
        )
        if not path:
            return
        try:
            with open(path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["Item Name", "Quantity", "Threshold",
                                 "Cost Price", "Sale Price", "Description"])
                for item in items:
                    writer.writerow(item)
            messagebox.showinfo("Export", f"Inventory exported to {path}")
        except Exception as e:
            messagebox.showerror("Export", f"Failed to export: {e}")

    def _export_excel(self):
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror(
                "Missing Dependency",
                "openpyxl is required for Excel export.\n\nInstall it with:\n  pip install openpyxl",
            )
            return
        items = self.app.inventory_service.get_all_items()
        if not items:
            messagebox.showwarning("Export", "No inventory data to export.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
            title="Export Inventory to Excel",
        )
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Inventory"
            ws.append(["Item Name", "Quantity", "Threshold",
                        "Cost Price", "Sale Price", "Description"])
            for item in items:
                ws.append(list(item))
            wb.save(path)
            messagebox.showinfo("Export", f"Inventory exported to {path}")
        except Exception as e:
            messagebox.showerror("Export", f"Failed to export: {e}")

    def _import_csv(self):
        path = filedialog.askopenfilename(
            filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")],
            title="Import Inventory from CSV",
        )
        if not path:
            return
        imported = skipped = 0
        try:
            with open(path, "r", encoding="utf-8") as f:
                for row in csv.DictReader(f):
                    try:
                        name = row.get("Item Name") or row.get("item_name")
                        qty  = int(row.get("Quantity")   or row.get("quantity")   or 0)
                        thr  = int(row.get("Threshold")  or row.get("threshold")  or 0)
                        cost = float(row.get("Cost Price") or row.get("cost_price") or 0)
                        sale = float(row.get("Sale Price") or row.get("sale_price") or 0)
                        desc = row.get("Description") or row.get("description") or ""
                        svc  = self.app.inventory_service
                        if svc.update_item(name, quantity=qty, threshold=thr,
                                           cost_price=cost, sale_price=sale,
                                           description=desc):
                            imported += 1
                        elif svc.add_item(name, qty, thr, cost, sale, desc):
                            imported += 1
                        else:
                            skipped += 1
                    except Exception:
                        skipped += 1
            self.refresh()
            messagebox.showinfo("Import", f"Imported: {imported}, Skipped: {skipped}")
        except Exception as e:
            messagebox.showerror("Import", f"Failed to import: {e}")

    def _import_excel(self):
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror(
                "Missing Dependency",
                "openpyxl is required for Excel import.\n\nInstall it with:\n  pip install openpyxl",
            )
            return
        path = filedialog.askopenfilename(
            filetypes=[("Excel Files", "*.xlsx *.xls"), ("All Files", "*.*")],
            title="Import Inventory from Excel",
        )
        if not path:
            return
        imported = skipped = 0
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            # Skip header row if first cell looks like a label
            start = 1 if rows and str(rows[0][0]).lower() in ("item name", "item_name", "name") else 0
            svc = self.app.inventory_service
            for row in rows[start:]:
                try:
                    name = str(row[0]).strip() if row[0] else ""
                    qty  = int(row[1] or 0)
                    thr  = int(row[2] or 0)
                    cost = float(row[3] or 0)
                    sale = float(row[4] or 0)
                    desc = str(row[5] or "")
                    if not name:
                        skipped += 1
                        continue
                    if svc.update_item(name, quantity=qty, threshold=thr,
                                       cost_price=cost, sale_price=sale, description=desc):
                        imported += 1
                    elif svc.add_item(name, qty, thr, cost, sale, desc):
                        imported += 1
                    else:
                        skipped += 1
                except Exception:
                    skipped += 1
            wb.close()
            self.refresh()
            messagebox.showinfo("Import", f"Imported: {imported}, Skipped: {skipped}")
        except Exception as e:
            messagebox.showerror("Import", f"Failed to import: {e}")
